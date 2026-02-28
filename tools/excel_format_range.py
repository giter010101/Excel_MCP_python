"""Tool: excel_format_range"""

from typing import Any, Optional
from fastmcp import FastMCP
from openpyxl.styles import Font, PatternFill, GradientFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.utils import get_column_letter
from excel_engine import open_workbook, save_workbook, parse_range, cell_name, _escape


def _apply_style(cell, style: dict):
    """Apply a style dict to an openpyxl cell."""
    if not style:
        return

    # Font
    font_cfg = style.get("font")
    if font_cfg:
        kw = {}
        if cell.font:
            kw.update({
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "vertAlign": cell.font.vertAlign,
                "underline": cell.font.underline,
                "strike": cell.font.strike,
                "color": cell.font.color
            })
        if "bold" in font_cfg: kw["bold"] = font_cfg["bold"]
        if "italic" in font_cfg: kw["italic"] = font_cfg["italic"]
        if "strike" in font_cfg: kw["strike"] = font_cfg["strike"]
        if "size" in font_cfg: kw["size"] = font_cfg["size"]
        if "color" in font_cfg: kw["color"] = font_cfg["color"].lstrip("#")
        if "underline" in font_cfg: kw["underline"] = font_cfg["underline"]
        cell.font = Font(**kw)

    # Alignment
    align_cfg = style.get("alignment")
    if align_cfg:
        kw = {}
        if cell.alignment:
            kw.update({
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text,
                "shrink_to_fit": cell.alignment.shrink_to_fit,
                "indent": cell.alignment.indent,
            })
        if "horizontal" in align_cfg: kw["horizontal"] = align_cfg["horizontal"]
        if "vertical" in align_cfg: kw["vertical"] = align_cfg["vertical"]
        if "wrapText" in align_cfg: kw["wrap_text"] = align_cfg["wrapText"]
        if "wrap_text" in align_cfg: kw["wrap_text"] = align_cfg["wrap_text"]
        cell.alignment = Alignment(**kw)

    # Fill
    fill_cfg = style.get("fill")
    if fill_cfg:
        if isinstance(fill_cfg, str):
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_cfg.lstrip("#"))
        else:
            fill_type = fill_cfg.get("type", "solid")
            if fill_type == "gradient":
                colors = fill_cfg.get("color", [])
                cell.fill = GradientFill(type="linear", stop=[c.lstrip("#") for c in colors])
            else:
                color = fill_cfg.get("color", "FFFFFF")
                if isinstance(color, list):
                    color = color[0] if color else "FFFFFF"
                color = color.lstrip("#")
                pattern = fill_cfg.get("pattern", "solid")
                if fill_cfg.get("type") == "pattern":
                    pass # Keep pattern if explicitly set via old type
                else:
                    pattern = "solid"
                cell.fill = PatternFill(fill_type=pattern, fgColor=color)

    # Border
    border_cfg = style.get("border", [])
    if border_cfg:
        sides = {}
        for b in border_cfg:
            btype = b.get("type", "")
            color = b.get("color", "#000000").lstrip("#")
            bstyle = b.get("style", "thin")
            side = Side(style=bstyle, color=color)
            if btype in ("left", "right", "top", "bottom", "diagonal"):
                sides[btype] = side
        cell.border = Border(**sides)

    # Number format
    num_fmt = style.get("numFmt")
    if num_fmt:
        cell.number_format = num_fmt

    decimal = style.get("decimalPlaces")
    if decimal is not None:
        # Build a simple number format with given decimal places
        cell.number_format = "0." + "0" * int(decimal)


def register_format_range(mcp: FastMCP):

    @mcp.tool(
        name="excel_format_range",
        description="Format cells in the Excel sheet with style information",
    )
    def excel_format_range(
        fileAbsolutePath: str,
        sheetName: str,
        range: str,
        style: Optional[dict] = None,
        styles: Optional[list[list[Optional[dict]]]] = None,
        autoFit: bool = False,
    ) -> str:
        """
        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Sheet name in the Excel file
            range: Range of cells (e.g. "A1:C3")
            style: A dictionary representing a single style to apply to the entire range.
            styles: 2D array of style objects for each cell. Use null to skip a cell.
            autoFit: Automatically adjust column widths based on contents in the range.
        """
        import builtins
        wb = open_workbook(fileAbsolutePath)
        if sheetName not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheetName!r}")
        ws = wb[sheetName]

        min_col, min_row, max_col, max_row = parse_range(range)
        range_row_size = max_row - min_row + 1
        range_col_size = max_col - min_col + 1

        if styles is not None:
            if len(styles) != range_row_size:
                raise ValueError(
                    f"Number of style rows ({len(styles)}) does not match range ({range_row_size})"
                )

        cells_processed = 0
        for i in builtins.range(range_row_size):
            for j in builtins.range(range_col_size):
                cell_style = None
                if style is not None:
                    cell_style = style
                elif styles is not None:
                    if len(styles[i]) != range_col_size:
                        raise ValueError(f"Number of style columns in row {i} ({len(styles[i])}) does not match range ({range_col_size})")
                    cell_style = styles[i][j]
                
                if cell_style is not None:
                    cell = ws.cell(row=min_row + i, column=min_col + j)
                    _apply_style(cell, cell_style)
                    cells_processed += 1

        if autoFit:
            for col in builtins.range(min_col, max_col + 1):
                col_letter = get_column_letter(col)
                max_length = 0
                for r in builtins.range(min_row, max_row + 1):
                    val = ws.cell(row=r, column=col).value
                    if val is not None:
                        max_length = max(max_length, len(str(val)))
                
                current_width = ws.column_dimensions[col_letter].width if ws.column_dimensions[col_letter].width else 8
                adjusted_width = max_length + 2
                if adjusted_width > current_width:
                    ws.column_dimensions[col_letter].width = adjusted_width

        save_workbook(wb, fileAbsolutePath)

        html = "<h2>Formatted Range</h2>\n"
        html += f"<p>Successfully applied styles to range {range} in sheet {_escape(sheetName)}</p>\n"
        html += "<h2>Metadata</h2>\n<ul>\n"
        html += "<li>backend: openpyxl</li>\n"
        html += f"<li>sheet name: {_escape(sheetName)}</li>\n"
        html += f"<li>formatted range: {range}</li>\n"
        html += f"<li>cells processed: {cells_processed}</li>\n"
        html += "</ul>\n<h2>Notice</h2>\n"
        html += "<p>Cell styles applied successfully.</p>\n"
        return html
