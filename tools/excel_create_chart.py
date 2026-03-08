"""Tool: excel_create_chart"""

_builtin_range = range  # save before parameter shadowing


from typing import Optional
from fastmcp import FastMCP
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, AreaChart,
    Reference, Series,
)
from excel_engine import open_workbook, save_workbook, parse_range, _escape, format_result


CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "scatter": ScatterChart,
    "area": AreaChart,
}


def register_create_chart(mcp: FastMCP):

    @mcp.tool(
        name="excel_create_chart",
        description="Create a chart in a worksheet",
    )
    def excel_create_chart(
        fileAbsolutePath: str,
        sheetName: str,
        range: str,
        chartType: str,
        title: Optional[str] = None,
        format: str = "json",
    ) -> str:
        """
        Create a chart from a data range.

        The range should be structured with:
        - First column: categories/labels for the X axis (text like names, dates, etc.)
        - Remaining columns: numeric data series (one series per column)
        - First row: headers (series names)

        Example layout for range "A1:C4":
            A         B        C
        1   Month     Sales    Costs
        2   Jan       100      80
        3   Feb       150      90
        4   Mar       200      110

        This creates a chart with "Month" as categories and "Sales"/"Costs" as two data series.

        For scatter charts: first column = X values, remaining columns = Y values.
        For pie charts: first column = labels, second column = values (only 1 data series).

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Name of the worksheet where the chart will be placed
            range: Data range including headers, structured as described above.
                   Can include sheet prefix like "Sheet1!A1:C10".
                   First column = categories, remaining columns = data series.
            chartType: Type of chart: line, bar, pie, scatter, area
            title: Optional title for the chart
            format: Output format — "json" (default) or "html"
        """
        chart_cls = CHART_TYPES.get(chartType.lower())
        if chart_cls is None:
            raise ValueError(
                f"Unknown chart type: {chartType!r}. "
                f"Supported: {', '.join(CHART_TYPES)}"
            )

        wb = open_workbook(fileAbsolutePath)
        if sheetName not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheetName!r}")
        ws = wb[sheetName]

        # Parse data range (may include sheet name prefix like "Sheet1!A1:B10")
        if "!" in range:
            data_sheet_name, data_range = range.split("!", 1)
            data_sheet_name = data_sheet_name.strip("'\"")
        else:
            data_sheet_name = sheetName
            data_range = range

        if data_sheet_name not in wb.sheetnames:
            raise ValueError(f"Data sheet not found: {data_sheet_name!r}")
        data_ws = wb[data_sheet_name]

        min_col, min_row, max_col, max_row = parse_range(data_range)
        num_data_cols = max_col - min_col  # columns after the first (categories)

        if num_data_cols < 1:
            raise ValueError(
                "Range must have at least 2 columns: "
                "1st column for categories/labels, remaining for data series."
            )

        chart = chart_cls()
        if title:
            chart.title = title
        chart.style = 10  # Clean default style

        # Categories = first column (skip header row)
        categories = Reference(
            data_ws,
            min_col=min_col,
            min_row=min_row + 1,
            max_row=max_row,
        )

        if chartType.lower() == "scatter":
            # Scatter: first col = X, remaining cols = Y series
            x_values = Reference(
                data_ws,
                min_col=min_col,
                min_row=min_row + 1,
                max_row=max_row,
            )
            for col_idx in _builtin_range(min_col + 1, max_col + 1):
                y_values = Reference(
                    data_ws,
                    min_col=col_idx,
                    min_row=min_row + 1,
                    max_row=max_row,
                )
                # Read series title from header
                series_title = data_ws.cell(row=min_row, column=col_idx).value
                series = Series(y_values, xvalues=x_values, title=str(series_title or f"Series {col_idx - min_col}"))
                chart.append(series)
        else:
            # All other chart types: data columns start after the categories column
            data_ref = Reference(
                data_ws,
                min_col=min_col + 1,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories)

        # Size
        chart.width = 15
        chart.height = 10

        # Anchor chart below data
        anchor_cell = f"A{max_row + 2}"
        ws.add_chart(chart, anchor_cell)

        save_workbook(wb, fileAbsolutePath)

        # Build response
        series_names = []
        for col_idx in _builtin_range(min_col + 1, max_col + 1):
            val = data_ws.cell(row=min_row, column=col_idx).value
            series_names.append(str(val) if val else f"Series {col_idx - min_col}")

        meta = {
            "backend": "openpyxl",
            "sheetName": sheetName,
            "chartType": chartType,
            "dataRange": range,
            "categoriesColumn": data_range.split(":")[0][0],
            "dataSeries": series_names,
            "dataRows": max_row - min_row,
            "anchor": anchor_cell,
        }
        if title:
            meta["title"] = title

        return format_result(
            action="Create Chart",
            message=f"{chartType} chart created in sheet '{sheetName}'.",
            metadata=meta,
            fmt=format,
        )
