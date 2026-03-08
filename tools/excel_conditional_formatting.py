"""Tool: excel_conditional_formatting (BETA)"""

from typing import Optional
from fastmcp import FastMCP
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    FormulaRule,
)
from excel_engine import open_workbook, save_workbook, _escape, format_result


def register_conditional_formatting(mcp: FastMCP):

    @mcp.tool(
        name="excel_conditional_formatting",
        description="Add conditional formatting rules to cells (color scales, data bars, icon sets, cell value rules, formula rules). (BETA)",
    )
    def excel_conditional_formatting(
        fileAbsolutePath: str,
        sheetName: str,
        range: str,
        ruleType: str,
        operator: Optional[str] = None,
        formula: Optional[list[str]] = None,
        values: Optional[list] = None,
        fillColor: Optional[str] = None,
        fontColor: Optional[str] = None,
        bold: Optional[bool] = None,
        colorScaleColors: Optional[list[str]] = None,
        dataBarColor: Optional[str] = None,
        iconStyle: Optional[str] = None,
        format: str = "json",
    ) -> str:
        """
        Add a conditional formatting rule to a range of cells.

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Sheet name in the Excel file
            range: Range to apply conditional formatting to (e.g. "B2:B100")
            ruleType: Type of conditional formatting rule. One of:
                - "cellIs"     : Format cells based on their value (uses operator + formula).
                - "colorScale" : Apply a 2-color or 3-color gradient (uses colorScaleColors).
                - "dataBar"    : Show data bars in cells (uses dataBarColor).
                - "iconSet"    : Show icons (arrows, traffic lights, etc.) (uses iconStyle + values).
                - "formula"    : Format cells based on a formula (uses formula).
            operator: For "cellIs" rule only. One of:
                "lessThan", "lessThanOrEqual", "greaterThan", "greaterThanOrEqual",
                "equal", "notEqual", "between", "notBetween"
            formula: Formula(s) for "cellIs" or "formula" rules.
                For "cellIs": e.g. ["0"] (compare to 0) or ["10","90"] (for between).
                For "formula": e.g. ["$A1>100"] (cells where formula is true get formatted).
            values: For "iconSet" rule: threshold values (e.g. [0, 33, 67] for 3 thresholds).
            fillColor: Background color for cellIs/formula rules (hex, e.g. "FFC7CE" for light red).
            fontColor: Font color for cellIs/formula rules (hex, e.g. "9C0006" for dark red).
            bold: Make text bold for cellIs/formula rules.
            colorScaleColors: List of 2 or 3 hex colors for colorScale rule.
                Example 2-color: ["F8696B", "63BE7B"] (red to green)
                Example 3-color: ["F8696B", "FFEB84", "63BE7B"] (red, yellow, green)
            dataBarColor: Hex color for the data bar (e.g. "638EC6" for blue).
            iconStyle: Icon set style name for iconSet rule. Common values:
                "3Arrows", "3ArrowsGray", "3Flags", "3TrafficLights1",
                "3TrafficLights2", "3Signs", "3Symbols", "3Symbols2",
                "4Arrows", "4ArrowsGray", "4RedToBlack", "4Rating",
                "4TrafficLights", "5Arrows", "5ArrowsGray", "5Rating", "5Quarters"
            format: Output format — "json" (default) or "html"
        """
        VALID_RULE_TYPES = {"cellIs", "colorScale", "dataBar", "iconSet", "formula"}
        if ruleType not in VALID_RULE_TYPES:
            raise ValueError(
                f"Invalid ruleType: {ruleType!r}. Must be one of: {', '.join(sorted(VALID_RULE_TYPES))}"
            )

        wb = open_workbook(fileAbsolutePath)
        if sheetName not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheetName!r}")
        ws = wb[sheetName]

        rule_desc = ruleType

        if ruleType == "cellIs":
            if not operator:
                raise ValueError("'operator' is required for cellIs rule")
            if not formula:
                raise ValueError("'formula' is required for cellIs rule")

            fmt_kwargs = {}
            if fillColor:
                fmt_kwargs["fill"] = PatternFill(
                    start_color=fillColor.lstrip("#"),
                    end_color=fillColor.lstrip("#"),
                    fill_type="solid",
                )
            if fontColor or bold is not None:
                font_kw = {}
                if fontColor:
                    font_kw["color"] = fontColor.lstrip("#")
                if bold is not None:
                    font_kw["bold"] = bold
                fmt_kwargs["font"] = Font(**font_kw)

            rule = CellIsRule(operator=operator, formula=formula, **fmt_kwargs)
            ws.conditional_formatting.add(range, rule)
            rule_desc = f"cellIs ({operator})"

        elif ruleType == "colorScale":
            if not colorScaleColors or len(colorScaleColors) < 2:
                raise ValueError("'colorScaleColors' must have 2 or 3 hex colors")

            colors = [c.lstrip("#") for c in colorScaleColors]
            if len(colors) == 2:
                rule = ColorScaleRule(
                    start_type="min", start_color=colors[0],
                    end_type="max", end_color=colors[1],
                )
            else:
                rule = ColorScaleRule(
                    start_type="min", start_color=colors[0],
                    mid_type="percentile", mid_value=50, mid_color=colors[1],
                    end_type="max", end_color=colors[2],
                )
            ws.conditional_formatting.add(range, rule)
            rule_desc = f"colorScale ({len(colors)} colors)"

        elif ruleType == "dataBar":
            bar_color = (dataBarColor or "638EC6").lstrip("#")
            rule = DataBarRule(
                start_type="min", end_type="max",
                color=bar_color,
            )
            ws.conditional_formatting.add(range, rule)
            rule_desc = "dataBar"

        elif ruleType == "iconSet":
            icon = iconStyle or "3TrafficLights1"
            icon_values = values or [0, 33, 67]
            rule = IconSetRule(
                icon_style=icon,
                type="percent",
                values=icon_values,
            )
            ws.conditional_formatting.add(range, rule)
            rule_desc = f"iconSet ({icon})"

        elif ruleType == "formula":
            if not formula:
                raise ValueError("'formula' is required for formula rule")

            fmt_kwargs = {}
            if fillColor:
                fmt_kwargs["fill"] = PatternFill(
                    start_color=fillColor.lstrip("#"),
                    end_color=fillColor.lstrip("#"),
                    fill_type="solid",
                )
            if fontColor or bold is not None:
                font_kw = {}
                if fontColor:
                    font_kw["color"] = fontColor.lstrip("#")
                if bold is not None:
                    font_kw["bold"] = bold
                fmt_kwargs["font"] = Font(**font_kw)

            rule = FormulaRule(formula=formula, **fmt_kwargs)
            ws.conditional_formatting.add(range, rule)
            rule_desc = "formula"

        save_workbook(wb, fileAbsolutePath)

        return format_result(
            action="Conditional Formatting",
            message=f"Rule '{rule_desc}' applied to range {range} in sheet '{sheetName}'.",
            metadata={
                "backend": "openpyxl",
                "sheetName": sheetName,
                "formattedRange": range,
                "ruleType": rule_desc,
            },
            fmt=format,
        )
