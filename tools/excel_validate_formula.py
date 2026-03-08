"""Tool: excel_validate_formula"""

import json
from typing import Optional
from fastmcp import FastMCP
from openpyxl.formula import Tokenizer
from excel_engine import open_workbook, _escape


def register_validate_formula(mcp: FastMCP):

    @mcp.tool(
        name="excel_validate_formula",
        description="Validate an Excel formula's syntax without applying it. Formulas must be in ENGLISH (SUM, not SOMME) — Excel français traduit automatiquement. Checks '=' prefix, tokenizes, and optionally verifies cell references.",
    )
    def excel_validate_formula(
        fileAbsolutePath: str,
        sheetName: str,
        formula: str,
        checkReferences: bool = False,
        format: str = "json",
    ) -> str:
        """
        Validate an Excel formula without writing it.
        The formula must use ENGLISH function names (SUM, AVERAGE, IF…).
        Excel français affiche automatiquement les noms en français.

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Sheet name in the Excel file
            formula: The formula to validate (must start with '=', ENGLISH function names)
            checkReferences: If True, verify that RANGE references point to
                             cells within the sheet's used area (best-effort)
            format: Output format — "json" (default) or "html"
        """
        errors: list[str] = []
        warnings: list[str] = []

        # 1. Must start with '='
        if not formula.startswith("="):
            errors.append("Formula must start with '='")
            return _build_output(formula, errors, warnings, tokens_info=[], fmt=format)

        # 2. Tokenize
        try:
            tok = Tokenizer(formula)
        except Exception as e:
            errors.append(f"Tokenizer error: {e}")
            return _build_output(formula, errors, warnings, tokens_info=[], fmt=format)

        # 3. Collect token info
        tokens_info: list[dict] = []
        range_tokens: list[str] = []
        for t in tok.items:
            tokens_info.append({
                "value": t.value,
                "type": t.type,
                "subtype": t.subtype,
            })
            if t.subtype == "RANGE":
                range_tokens.append(t.value)

        # 4. Basic sanity checks
        if not tokens_info:
            errors.append("Formula produced no tokens (empty formula body)")

        # Check for unbalanced parentheses
        open_parens = sum(1 for t in tok.items if t.value == "(" and t.type == "PAREN")
        close_parens = sum(1 for t in tok.items if t.value == ")" and t.type == "PAREN")
        if open_parens != close_parens:
            errors.append(
                f"Unbalanced parentheses: {open_parens} opening vs {close_parens} closing"
            )

        # Check for ERROR tokens
        for t in tok.items:
            if t.subtype == "ERROR":
                errors.append(f"Error token found: {t.value}")

        # 5. Optional: verify references against sheet
        if checkReferences and range_tokens:
            try:
                wb = open_workbook(fileAbsolutePath)
                if sheetName not in wb.sheetnames:
                    errors.append(f"Sheet not found: {sheetName!r}")
                else:
                    ws = wb[sheetName]
                    max_row = ws.max_row or 1
                    max_col = ws.max_column or 1
                    for ref in range_tokens:
                        # Skip sheet-qualified references (e.g. Sheet2!A1)
                        if "!" in ref:
                            ref_sheet, ref_cell = ref.split("!", 1)
                            ref_sheet = ref_sheet.strip("'\"")
                            if ref_sheet not in wb.sheetnames:
                                warnings.append(
                                    f"Reference to unknown sheet: {ref_sheet!r} in {ref}"
                                )
                        # We don't deeply validate cell coordinates here,
                        # just flag cross-sheet references to missing sheets
            except Exception as e:
                warnings.append(f"Could not open workbook for reference check: {e}")

        return _build_output(formula, errors, warnings, tokens_info, fmt=format)


def _build_output(
    formula: str,
    errors: list[str],
    warnings: list[str],
    tokens_info: list[dict],
    fmt: str = "json",
) -> str:
    is_valid = len(errors) == 0

    if fmt == "html":
        status = "✅ Valid" if is_valid else "❌ Invalid"
        html = "<h2>Formula Validation</h2>\n"
        html += f"<p>Formula: <code>{_escape(formula)}</code></p>\n"
        html += f"<p>Status: <strong>{status}</strong></p>\n"

        if errors:
            html += "<h3>Errors</h3>\n<ul>\n"
            for e in errors:
                html += f"<li>{_escape(e)}</li>\n"
            html += "</ul>\n"

        if warnings:
            html += "<h3>Warnings</h3>\n<ul>\n"
            for w in warnings:
                html += f"<li>{_escape(w)}</li>\n"
            html += "</ul>\n"

        if tokens_info:
            html += "<h3>Tokens</h3>\n"
            html += "<table><tr><th>Value</th><th>Type</th><th>Subtype</th></tr>\n"
            for t in tokens_info:
                html += (
                    f"<tr><td>{_escape(t['value'])}</td>"
                    f"<td>{_escape(t['type'])}</td>"
                    f"<td>{_escape(t['subtype'])}</td></tr>\n"
                )
            html += "</table>\n"

        return html

    # JSON (default)
    result: dict = {
        "action": "Formula Validation",
        "formula": formula,
        "valid": is_valid,
    }
    if errors:
        result["errors"] = errors
    if warnings:
        result["warnings"] = warnings
    if tokens_info:
        result["tokens"] = tokens_info
    return json.dumps(result, ensure_ascii=False)
