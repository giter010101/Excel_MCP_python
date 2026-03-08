"""Tool: excel_copy_range — Copy a range of cells (values + styles + formulas) to another location."""

from typing import Optional
from copy import copy
from fastmcp import FastMCP
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from excel_engine import open_workbook, save_workbook, parse_range, cell_name, _escape, format_result


def register_copy_range(mcp: FastMCP):

    @mcp.tool(
        name="excel_copy_range",
        description="Copy a range of cells (values, formulas, and styles) to another location, in the same sheet or a different sheet.",
    )
    def excel_copy_range(
        fileAbsolutePath: str,
        sheetName: str,
        sourceRange: str,
        targetStartCell: str,
        targetSheetName: Optional[str] = None,
        translateFormulas: bool = True,
        format: str = "json",
    ) -> str:
        """
        Copy a range of cells to another location within the same workbook.

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Source sheet name
            sourceRange: Range to copy (e.g. "A1:D10")
            targetStartCell: Top-left cell of the destination (e.g. "F1")
            targetSheetName: Destination sheet name. Defaults to the source sheet.
            translateFormulas: If True, formulas are adjusted (translated) to the
                               new cell positions. If False, formulas are copied verbatim.
            format: Output format — "json" (default) or "html"
        """
        wb = open_workbook(fileAbsolutePath)

        if sheetName not in wb.sheetnames:
            raise ValueError(f"Source sheet not found: {sheetName!r}")

        target_sheet = targetSheetName or sheetName
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"Target sheet not found: {target_sheet!r}")

        src_ws = wb[sheetName]
        tgt_ws = wb[target_sheet]

        # Parse source range
        src_min_col, src_min_row, src_max_col, src_max_row = parse_range(sourceRange)

        # Parse target start cell (reuse parse_range with a single-cell range)
        from openpyxl.utils import coordinate_to_tuple
        tgt_start_row, tgt_start_col = coordinate_to_tuple(targetStartCell)

        cells_copied = 0
        for row_offset in range(src_max_row - src_min_row + 1):
            for col_offset in range(src_max_col - src_min_col + 1):
                src_row = src_min_row + row_offset
                src_col = src_min_col + col_offset
                tgt_row = tgt_start_row + row_offset
                tgt_col = tgt_start_col + col_offset

                src_cell = src_ws.cell(row=src_row, column=src_col)
                tgt_cell = tgt_ws.cell(row=tgt_row, column=tgt_col)

                # --- Copy value / formula ---
                value = src_cell.value
                if (
                    translateFormulas
                    and isinstance(value, str)
                    and value.startswith("=")
                ):
                    try:
                        origin = f"{get_column_letter(src_col)}{src_row}"
                        dest = f"{get_column_letter(tgt_col)}{tgt_row}"
                        value = Translator(value, origin=origin).translate_formula(dest)
                    except Exception:
                        pass  # Keep original formula if translation fails
                tgt_cell.value = value

                # --- Copy styles ---
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.alignment = copy(src_cell.alignment)
                    tgt_cell.number_format = src_cell.number_format
                    tgt_cell.protection = copy(src_cell.protection)

                cells_copied += 1

        save_workbook(wb, fileAbsolutePath)

        tgt_end_row = tgt_start_row + (src_max_row - src_min_row)
        tgt_end_col = tgt_start_col + (src_max_col - src_min_col)
        target_range = f"{cell_name(tgt_start_col, tgt_start_row)}:{cell_name(tgt_end_col, tgt_end_row)}"

        return format_result(
            action="Copy Range",
            message=f"Copied {sourceRange} from '{sheetName}' to {target_range} in '{target_sheet}'.",
            metadata={
                "backend": "openpyxl",
                "sheetName": sheetName,
                "cellsCopied": cells_copied,
                "formulasTranslated": translateFormulas,
            },
            fmt=format,
        )
