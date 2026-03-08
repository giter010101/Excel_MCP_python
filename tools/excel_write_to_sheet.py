"""Tool: excel_write_to_sheet"""

import json
from typing import Any, Optional
from fastmcp import FastMCP
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from excel_engine import (
    open_workbook, save_workbook, parse_range,
    build_html_table, build_json_table, cell_name, _escape
)


def register_write_to_sheet(mcp: FastMCP):

    @mcp.tool(
        name="excel_write_to_sheet",
        description="Write values to the Excel sheet",
    )
    def excel_write_to_sheet(
        fileAbsolutePath: str,
        sheetName: str,
        values: list[list[Any]],
        newSheet: bool = False,
        range: Optional[str] = None,
        startCell: Optional[str] = None,
        append: bool = False,
        format: str = "json",
    ) -> str:
        """
        BEST PRACTICE - USE FORMULAS, NOT HARD-CODED VALUES:
        When writing data that involves ANY calculation (totals, sums, averages,
        percentages, differences, running totals, etc.), ALWAYS use Excel formulas
        (e.g. "=SUM(B2:B10)") instead of computing the result yourself and writing
        a hard-coded number. This ensures the spreadsheet remains dynamic and
        recalculable by the user. Only use literal values for raw input data.

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Sheet name in the Excel file
            values: 2D array of values to write. Formulas start with "=".
                    IMPORTANT - FORMULA RULES (SERVER POUR EXCEL FRANÇAIS):
                    1. TOUJOURS écrire les formules en ANGLAIS (SUM, AVERAGE, IF, VLOOKUP…).
                       C'est le format interne .xlsx. Excel français les affiche automatiquement
                       en français à l'utilisateur (SUM → SOMME, AVERAGE → MOYENNE, etc.).
                    2. NE JAMAIS utiliser les noms français (SOMME, MOYENNE, SI, RECHERCHEV…)
                       car openpyxl les stockera tels quels et Excel affichera des erreurs.
                    3. NE JAMAIS utiliser de fonctions avec un point (.) dans leur nom.
                       openpyxl n'ajoute pas le préfixe _xlfn. requis → erreurs #NOM? dans Excel.
                       INTERDIT (avec point): NORM.DIST, NORM.S.DIST, T.DIST, BINOM.DIST, etc.
                       UTILISER À LA PLACE (legacy): NORMDIST, NORMSDIST, TDIST, BINOMDIST, etc.
                    4. Utiliser le point-virgule (;) comme séparateur n'est PAS nécessaire.
                       Écrire =SUM(A1,A2) avec des virgules. Excel français traduit automatiquement.
                    5. Exemples: "=SUM(A1:A10)", "=AVERAGE(B2:B5)", "=IF(A1>0,A1,0)",
                       "=VLOOKUP(A1,Data!A:B,2,FALSE)", "=COUNTIF(A:A,\">0\")"
            newSheet: Create a new sheet if true, otherwise write to the existing sheet
            range: Range of cells (e.g. "A1:C10"). Not needed if startCell or append is used.
            startCell: Top-left cell to start writing from (e.g. "A1"). Will automatically calculate bounds.
            append: If True, writes data at the first empty row at the bottom of the sheet.
            format: Output format — "json" (default, compact) or "html" (legacy verbose)
        """
        import os
        from excel_engine import create_workbook
        
        if not os.path.exists(fileAbsolutePath):
            wb = create_workbook(fileAbsolutePath)
        else:
            wb = open_workbook(fileAbsolutePath)

        if newSheet:
            if sheetName in wb.sheetnames:
                pass # Already exists, we can still use it or raise? Let's use it or just pass
            else:
                wb.create_sheet(sheetName)

        if sheetName not in wb.sheetnames:
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
                wb["Sheet"].title = sheetName
            else:
                wb.create_sheet(sheetName)

        ws = wb[sheetName]
        
        if append:
            start_r = ws.max_row
            # If sheet is totally empty, max_row is 1, but let's check if A1 has value
            if start_r == 1 and ws.cell(row=1, column=1).value is None:
                start_r = 1
            else:
                start_r += 1
            startCell = f"A{start_r}"

        if not range and not startCell:
            raise ValueError("You must provide either 'range', 'startCell', or set 'append'=True")

        if startCell:
            min_row, min_col = coordinate_to_tuple(startCell)
            range_row_size = len(values)
            range_col_size = max(len(row) for row in values) if values else 0
            max_row = min_row + range_row_size - 1
            max_col = min_col + range_col_size - 1
            range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        else:
            min_col, min_row, max_col, max_row = parse_range(range)
            range_row_size = max_row - min_row + 1
            range_col_size = max_col - min_col + 1
            if len(values) != range_row_size:
                raise ValueError(
                    f"Number of rows in data ({len(values)}) does not match range size ({range_row_size})"
                )

        wrote_formula = False

        for i, row in enumerate(values):
            # For startCell/append, we tolerate rows of different lengths.
            if startCell is None and len(row) != range_col_size:
                raise ValueError(
                    f"Number of columns in row {i} ({len(row)}) does not match range size ({range_col_size})"
                )
            for j, val in enumerate(row):
                cell = ws.cell(row=min_row + i, column=min_col + j)
                if isinstance(val, str) and val.startswith("="):
                    cell.value = val
                    wrote_formula = True
                else:
                    cell.value = val

        save_workbook(wb, fileAbsolutePath)

        # Reload to show written result
        wb2 = open_workbook(fileAbsolutePath)
        ws2 = wb2[sheetName]

        if format == "html":
            table_html = build_html_table(ws2, min_col, min_row, max_col, max_row, wrote_formula)
            wb2.close()
            html = "<h2>Written Sheet</h2>\n"
            html += table_html + "\n"
            html += "<h2>Metadata</h2>\n<ul>\n"
            html += "<li>backend: openpyxl</li>\n"
            html += f"<li>sheet name: {_escape(sheetName)}</li>\n"
            html += f"<li>read range: {range}</li>\n"
            html += "</ul>\n<h2>Notice</h2>\n"
            html += "<p>Values wrote successfully.</p>\n"
            return html
        else:
            columns, rows = build_json_table(ws2, min_col, min_row, max_col, max_row, wrote_formula)
            wb2.close()
            result = {
                "action": "write_to_sheet",
                "message": "Values written successfully.",
                "sheet": sheetName,
                "range": range,
                "columns": columns,
                "rows": rows,
            }
            return json.dumps(result, ensure_ascii=False, default=str)
