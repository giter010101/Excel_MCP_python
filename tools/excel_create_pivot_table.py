"""Tool: excel_create_pivot_table — uses pandas to compute the pivot and openpyxl to write it."""

from typing import Optional
from fastmcp import FastMCP
from excel_engine import _escape


def register_create_pivot_table(mcp: FastMCP):

    @mcp.tool(
        name="excel_create_pivot_table",
        description=(
            "Create a pivot table from a data range. The pivot is computed with pandas "
            "and written to a new (or existing) worksheet. openpyxl does not support "
            "native Excel PivotTable objects, so the result is a flat table of aggregated data."
        ),
    )
    def excel_create_pivot_table(
        fileAbsolutePath: str,
        sheetName: str,
        dataRange: str,
        targetSheet: str,
        rows: list[str],
        values: list[str],
        columns: Optional[list[str]] = None,
        aggFunc: str = "sum",
        targetCell: str = "A1",
    ) -> str:
        """
        Compute a pivot table from source data and write the result.

        Args:
            fileAbsolutePath: Absolute path to the Excel file
            sheetName: Sheet containing the source data
            dataRange: Range of source data including headers (e.g. "A1:E20")
            targetSheet: Name of the worksheet to write the pivot result into.
                         Created automatically if it does not exist.
            rows: Column header names to use as pivot rows (index).
            values: Column header names to aggregate.
            columns: Optional column header names to use as pivot columns.
            aggFunc: Aggregation function: "sum", "mean", "count", "min", "max", "median".
                     Default is "sum".
            targetCell: Top-left cell to start writing pivot result (default "A1").
        """
        try:
            import pandas as pd
        except ImportError:
            raise RuntimeError(
                "pandas is required for pivot table creation. "
                "Install it with: pip install pandas"
            )
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import coordinate_to_tuple
        from excel_engine import parse_range

        VALID_AGG = {"sum", "mean", "count", "min", "max", "median"}
        if aggFunc not in VALID_AGG:
            raise ValueError(
                f"Invalid aggFunc: {aggFunc!r}. Must be one of: {', '.join(sorted(VALID_AGG))}"
            )

        # 1. Read source data into pandas DataFrame
        wb = load_workbook(fileAbsolutePath)
        if sheetName not in wb.sheetnames:
            raise ValueError(f"Source sheet not found: {sheetName!r}")
        ws = wb[sheetName]

        min_col, min_row, max_col, max_row = parse_range(dataRange)

        # Extract headers from first row
        headers = []
        for col in range(min_col, max_col + 1):
            val = ws.cell(row=min_row, column=col).value
            headers.append(str(val) if val is not None else f"Col{col}")

        # Extract data rows
        data_rows = []
        for r in range(min_row + 1, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                row_data.append(ws.cell(row=r, column=c).value)
            data_rows.append(row_data)

        df = pd.DataFrame(data_rows, columns=headers)

        # Validate column names
        for col_name in rows:
            if col_name not in df.columns:
                raise ValueError(
                    f"Row field '{col_name}' not found in data headers: {list(df.columns)}"
                )
        for col_name in values:
            if col_name not in df.columns:
                raise ValueError(
                    f"Value field '{col_name}' not found in data headers: {list(df.columns)}"
                )
        if columns:
            for col_name in columns:
                if col_name not in df.columns:
                    raise ValueError(
                        f"Column field '{col_name}' not found in data headers: {list(df.columns)}"
                    )

        # 2. Compute pivot
        pivot_df = pd.pivot_table(
            df,
            values=values,
            index=rows,
            columns=columns or [],
            aggfunc=aggFunc,
            fill_value=0,
        )

        # Flatten MultiIndex columns if present
        if isinstance(pivot_df.columns, pd.MultiIndex):
            pivot_df.columns = [
                " | ".join(str(c) for c in col_tuple).strip(" | ")
                for col_tuple in pivot_df.columns
            ]

        pivot_df = pivot_df.reset_index()

        # 3. Write to target sheet
        if targetSheet not in wb.sheetnames:
            wb.create_sheet(targetSheet)
        tgt_ws = wb[targetSheet]

        tgt_start_row, tgt_start_col = coordinate_to_tuple(targetCell)

        for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = tgt_ws.cell(
                    row=tgt_start_row + r_idx,
                    column=tgt_start_col + c_idx,
                    value=value,
                )

        wb.save(fileAbsolutePath)

        result_rows = len(pivot_df)
        result_cols = len(pivot_df.columns)

        html = "<h2>Pivot Table Created</h2>\n"
        html += f"<p>Pivot table computed from '{_escape(sheetName)}' range {dataRange} "
        html += f"and written to sheet '{_escape(targetSheet)}' starting at {targetCell}.</p>\n"
        html += "<h2>Details</h2>\n<ul>\n"
        html += f"<li>rows (index): {', '.join(rows)}</li>\n"
        html += f"<li>values: {', '.join(values)}</li>\n"
        html += f"<li>columns: {', '.join(columns) if columns else '(none)'}</li>\n"
        html += f"<li>aggregation: {aggFunc}</li>\n"
        html += f"<li>result size: {result_rows} rows × {result_cols} columns</li>\n"
        html += "</ul>\n"
        html += "<h2>Metadata</h2>\n<ul>\n"
        html += "<li>backend: openpyxl + pandas</li>\n"
        html += f"<li>source sheet: {_escape(sheetName)}</li>\n"
        html += f"<li>target sheet: {_escape(targetSheet)}</li>\n"
        html += "</ul>\n"
        return html
