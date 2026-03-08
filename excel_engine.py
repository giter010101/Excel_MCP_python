"""
Excel Engine - wraps openpyxl for all Excel file operations.
Equivalent to the Go internal/excel package.
"""

import os
import re
from typing import Any, Optional
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    coordinate_to_tuple,
)
from openpyxl.styles import (
    Font, PatternFill, GradientFill, Border, Side, Alignment, numbers
)
from openpyxl.utils.cell import range_boundaries  # (min_col, min_row, max_col, max_row)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

PAGING_CELLS_LIMIT = int(os.environ.get("EXCEL_MCP_PAGING_CELLS_LIMIT", "2000"))

_RANGE_RE = re.compile(
    r"^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$"
)


def parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Return (min_col, min_row, max_col, max_row) as 1-based ints."""
    m = _RANGE_RE.match(range_str.strip())
    if not m:
        raise ValueError(f"Invalid range: {range_str!r}")
    min_col = column_index_from_string(m.group(1))
    min_row = int(m.group(2))
    max_col = column_index_from_string(m.group(3))
    max_row = int(m.group(4))
    return min_col, min_row, max_col, max_row


def col_to_letter(col: int) -> str:
    return get_column_letter(col)


def cell_name(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


def get_used_range(ws) -> Optional[str]:
    """Return the used range of a worksheet as a string like 'A1:D10'."""
    if ws.max_row is None or ws.max_column is None:
        return None
    if ws.max_row == 0 or ws.max_column == 0:
        return None
    # openpyxl dimensions
    min_r = ws.min_row or 1
    min_c = ws.min_column or 1
    max_r = ws.max_row
    max_c = ws.max_column
    return f"{cell_name(min_c, min_r)}:{cell_name(max_c, max_r)}"


def get_paging_ranges(ws, limit: int = PAGING_CELLS_LIMIT) -> list[str]:
    """Split the used range into pages of at most `limit` cells."""
    used = get_used_range(ws)
    if not used:
        return []
    min_col, min_row, max_col, max_row = parse_range(used)
    total_cols = max_col - min_col + 1
    rows_per_page = max(1, limit // total_cols)

    pages = []
    r = min_row
    while r <= max_row:
        end_r = min(r + rows_per_page - 1, max_row)
        pages.append(f"{cell_name(min_col, r)}:{cell_name(max_col, end_r)}")
        r = end_r + 1
    return pages


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def open_workbook(path: str, read_only: bool = False) -> Workbook:
    if not os.path.isabs(path):
        raise ValueError(f"Path must be absolute: {path!r}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path!r}")
    return load_workbook(path, data_only=read_only, keep_vba=True if path.endswith((".xlsm", ".xlsb")) else False)


def save_workbook(wb: Workbook, path: str) -> None:
    wb.save(path)
    wb.close()


def create_workbook(path: str) -> Workbook:
    if not os.path.isabs(path):
        raise ValueError(f"Path must be absolute: {path!r}")
    wb = Workbook()
    wb.save(path)
    return wb


# ---------------------------------------------------------------------------
# HTML table builder (mirrors Go createHTMLTable)
# ---------------------------------------------------------------------------

def _escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
    )


def build_html_table(
    ws,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    show_formula: bool = False,
) -> str:
    rows_html = []
    # Header row with column letters
    header_cells = ["<th></th>"]
    for c in range(min_col, max_col + 1):
        header_cells.append(f"<th>{get_column_letter(c)}</th>")
    rows_html.append("<tr>" + "".join(header_cells) + "</tr>")

    for r in range(min_row, max_row + 1):
        cells = [f"<th>{r}</th>"]
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if show_formula and cell.data_type == "f":
                value = cell.value or ""
            else:
                v = cell.value
                value = "" if v is None else str(v)
            cells.append(f"<td>{_escape(value).replace(chr(10), '<br>')}</td>")
        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    return "<table>\n" + "\n".join(rows_html) + "\n</table>"


def read_sheet_html(
    wb: Workbook,
    sheet_name: str,
    range_str: Optional[str],
    show_formula: bool,
    show_style: bool,
) -> dict:
    """
    Returns a dict with keys: html, read_range, next_range, sheet_name.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name!r}")
    ws = wb[sheet_name]

    pages = get_paging_ranges(ws)
    if not pages:
        raise ValueError("No range available to read")

    current_range = range_str if range_str else pages[0]

    # find next range
    next_range = ""
    try:
        idx = pages.index(current_range)
        if idx + 1 < len(pages):
            next_range = pages[idx + 1]
    except ValueError:
        pass  # custom range, no next

    min_col, min_row, max_col, max_row = parse_range(current_range)

    table_html = build_html_table(ws, min_col, min_row, max_col, max_row, show_formula)

    result = f"<h2>Read Sheet</h2>\n{table_html}\n"
    result += "<h2>Metadata</h2>\n<ul>\n"
    result += f"<li>backend: openpyxl</li>\n"
    result += f"<li>sheet name: {_escape(sheet_name)}</li>\n"
    result += f"<li>read range: {current_range}</li>\n"
    result += "</ul>\n<h2>Notice</h2>\n"
    if next_range:
        result += "<p>This sheet has more ranges.</p>\n"
        result += "<p>To read the next range, specify 'range' argument as follows.</p>\n"
        result += f'<code>{{ "range": "{next_range}" }}</code>\n'
    else:
        result += "<p>This is the last range or no more ranges available.</p>\n"

    return {"html": result, "read_range": current_range, "next_range": next_range}


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def build_json_table(
    ws,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    show_formula: bool = False,
) -> tuple[list, dict]:
    """
    Return (columns, rows) where:
    - columns: list of column letters e.g. ["A", "B", "C"]
    - rows: dict mapping str(excel_row_number) -> list of values
            e.g. {"1": ["Name", "Price"], "2": ["Laptop", 999]}
    Keys are the ACTUAL Excel row numbers so no index arithmetic is needed.
    """
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    rows: dict = {}
    for r in range(min_row, max_row + 1):
        row_data = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if show_formula and cell.data_type == "f":
                row_data.append(cell.value)
            else:
                row_data.append(cell.value)
        rows[str(r)] = row_data
    return columns, rows


def read_sheet_json(
    wb: Workbook,
    sheet_name: str,
    range_str: Optional[str],
    show_formula: bool,
    show_style: bool,
) -> dict:
    """
    Returns a dict with keys: sheet, range, columns, rows, nextRange.
    - columns: list of column letters in order (e.g. ["A","B","C"])
    - rows: dict mapping Excel row number (as string) to list of values
            e.g. {"10": ["Laptop", 999, "=SUM(B10:C10)"]}
    Row keys ARE the exact Excel row numbers — no index arithmetic needed.
    To get cell B10: rows["10"][columns.index("B")]
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name!r}")
    ws = wb[sheet_name]

    pages = get_paging_ranges(ws)
    if not pages:
        raise ValueError("No range available to read")

    current_range = range_str if range_str else pages[0]

    next_range = ""
    try:
        idx = pages.index(current_range)
        if idx + 1 < len(pages):
            next_range = pages[idx + 1]
    except ValueError:
        pass

    min_col, min_row, max_col, max_row = parse_range(current_range)
    columns, rows = build_json_table(ws, min_col, min_row, max_col, max_row, show_formula)

    result: dict = {
        "sheet": sheet_name,
        "range": current_range,
        "columns": columns,
        "rows": rows,
    }
    if next_range:
        result["nextRange"] = next_range

    return result


def format_result(
    action: str,
    message: str,
    metadata: dict,
    fmt: str = "json",
) -> str:
    """
    Return a confirmation string in `fmt` format ('json' or 'html').

    - JSON: compact dict with action, message, plus all metadata keys.
    - HTML: the classic <h2>/<p>/<ul> pattern.
    """
    import json as _json

    if fmt == "html":
        html = f"<h2>{_escape(action)}</h2>\n"
        html += f"<p>{_escape(message)}</p>\n"
        if metadata:
            html += "<h2>Metadata</h2>\n<ul>\n"
            for k, v in metadata.items():
                html += f"<li>{_escape(str(k))}: {_escape(str(v))}</li>\n"
            html += "</ul>\n"
        return html

    # JSON (default)
    payload: dict = {"action": action, "message": message}
    payload.update(metadata)
    return _json.dumps(payload, ensure_ascii=False)
