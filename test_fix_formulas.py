import builtins
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.load_workbook('c:\\Users\\natha\\Documents\\MCP_server\\Test\\test_mcp.xlsx', data_only=False)
ws = wb['Budget PME']

# Re-write the formulas directly to be absolutely sure
formulas = {
    'G4': '=SUM(C4:F4)', 'G5': '=SUM(C5:F5)',
    'C6': '=SUM(C4:C5)', 'D6': '=SUM(D4:D5)', 'E6': '=SUM(E4:E5)', 'F6': '=SUM(F4:F5)', 'G6': '=SUM(G4:G5)',
    'G8': '=SUM(C8:F8)', 'G9': '=SUM(C9:F9)', 'G10': '=SUM(C10:F10)', 'G11': '=SUM(C11:F11)',
    'C12': '=SUM(C8:C11)', 'D12': '=SUM(D8:D11)', 'E12': '=SUM(E8:E11)', 'F12': '=SUM(F8:F11)', 'G12': '=SUM(G8:G11)',
    'C13': '=C6+C12', 'D13': '=D6+D12', 'E13': '=E6+E12', 'F13': '=F6+F12', 'G13': '=G6+G12'
}
for k, v in formulas.items():
    ws[k].value = v

# Formatting
for r in ws['C4:G13']:
    for c in r:
        c.number_format = '# ##0 €'

fill_header = PatternFill(fill_type='solid', fgColor='1F4E78')
font_header = Font(bold=True, color='FFFFFF', size=12)
align_center = Alignment(horizontal='center', vertical='center')
for r in ws['B2:G2']:
    for c in r:
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center

fill_sub1 = PatternFill(fill_type='solid', fgColor='D9D9D9')
font_sub1 = Font(bold=True, italic=True)
for r in list(ws['B3:G3']) + list(ws['B7:G7']):
    for c in r:
        c.fill = fill_sub1
        c.font = font_sub1

fill_tot_rev = PatternFill(fill_type='solid', fgColor='E2EFDA')
font_tot_rev = Font(bold=True, color='375623')
for r in ws['B6:G6']:
    for c in r:
        c.fill = fill_tot_rev
        c.font = font_tot_rev

fill_tot_dep = PatternFill(fill_type='solid', fgColor='FCE4D6')
font_tot_dep = Font(bold=True, color='C00000')
for r in ws['B12:G12']:
    for c in r:
        c.fill = fill_tot_dep
        c.font = font_tot_dep

fill_net = PatternFill(fill_type='solid', fgColor='FFF2CC')
font_net = Font(bold=True, size=13)
for r in ws['B13:G13']:
    for c in r:
        c.fill = fill_net
        c.font = font_net

# Autofit
from openpyxl.utils import get_column_letter
for col in builtins.range(2, 8):
    col_letter = get_column_letter(col)
    max_length = 0
    for row in builtins.range(2, 14):
        val = ws.cell(row=row, column=col).value
        # For autofit, openpyxl treats formulas as just strings, so len might be small or exactly the formula text.
        if val is not None:
            max_length = max(max_length, len(str(val)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save('c:\\Users\\natha\\Documents\\MCP_server\\Test\\test_mcp.xlsx')
